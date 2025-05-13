from flask import Blueprint, render_template, request, flash, redirect, url_for, make_response, Response, current_app, send_file
from flask_login import login_required, current_user
from flask_mail import Message
from extensions import db, mail
from functools import wraps
import click, logging, random, string
from models import User, UserRole, AuditLog, Notification, Cuti, CutiStatus, JenisCuti, db
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta
import pandas as pd
from enum import Enum
from io import BytesIO
from fpdf import FPDF
from sqlalchemy import extract, or_, and_, func, text
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from services.email import send_welcome_email, send_admin_verification_email, send_admin_status_email, send_admin_password_reset_email, EmailSendError
from services.report_generator import generate_excel_report, generate_pdf_report
from utils.decorators import role_required
from utils.tokens import generate_email_token
from utils.validators import validate_email, validate_password

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__)

# ===================== CLI COMMAND =====================
@admin_bp.cli.command("create-superadmin")
@click.argument("username")
@click.argument("email")
@click.argument("password")
def create_superadmin(username, email, password):
    """Membuat akun superadmin baru"""
    if User.query.filter_by(role='superadmin').first():
        print("Superadmin sudah ada!")
        return

    superadmin = User(
        username=username,
        email=email,
        password=generate_password_hash(password, method='pbkdf2:sha256'),
        role='superadmin',
        email_verified=True,
        jabatan='Super Administrator'
    )
    
    db.session.add(superadmin)
    db.session.commit()
    print(f"Superadmin {username} berhasil dibuat!")

# ===================== SUPERADMIN ROUTES =====================
@admin_bp.route('/superadmin/dashboard')
@login_required
@role_required(UserRole.SUPERADMIN)
def superadmin_dashboard():
    try:
        # Get recent audit logs (using created_at instead of timestamp)
        audit_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(20).all()
        
        # Get unread notifications
        notifications = Notification.query.filter_by(
            user_id=current_user.id,
            is_read=False
        ).order_by(
            Notification.created_at.desc()
        ).limit(10).all()
        
        # Get stats
        total_admin = User.query.filter_by(role=UserRole.ADMIN, is_active=True).count()
        total_atasan = User.query.filter_by(role=UserRole.ATASAN, is_active=True).count()
        total_pegawai = User.query.filter_by(role=UserRole.PEGAWAI, is_active=True).count()
        
        return render_template(
            'superadmin/dashboard.html',
            audit_logs=audit_logs,
            notifications=notifications,
            total_admin=total_admin,
            total_atasan=total_atasan,
            total_pegawai=total_pegawai
        )
        
    except Exception as e:
        current_app.logger.error(f"Dashboard error: {str(e)}", exc_info=True)
        flash('Terjadi kesalahan saat memuat dashboard', 'error')
        return redirect(url_for('admin.manajemen_admin'))

@admin_bp.route('/superadmin/manajemen-admin', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def manajemen_admin():
    if request.method == 'POST':
        return handle_admin_creation()
        
    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '').strip()
    
    # Hanya tampilkan admin yang aktif (is_active=True)
    query = User.query.filter_by(role=UserRole.ADMIN, is_active=True)
    
    if search_query:
        query = query.filter(
            (User.username.ilike(f'%{search_query}%')) |
            (User.email.ilike(f'%{search_query}%')) |
            (User.full_name.ilike(f'%{search_query}%'))
        )
    
    admins = query.order_by(User.created_at.desc()).paginate(
        page=page, 
        per_page=10, 
        error_out=False
    )
    
    return render_template('superadmin/manajemen_admin.html', 
                         admins=admins,
                         search_query=search_query)

def handle_admin_creation():
    """Handle admin creation form submission"""
    form_data = {
        'nip': request.form.get('nip', '').strip(),
        'username': request.form.get('username', '').strip(),
        'full_name': request.form.get('full_name', '').strip(),
        'email': request.form.get('email', '').lower().strip(),
        'password': request.form.get('password', ''),
        'phone': request.form.get('phone', '').strip()
    }
    
    # Validate form
    errors = validate_admin_form(form_data)
    if errors:
        for error in errors:
            flash(error, 'error')
        return redirect(url_for('admin.manajemen_admin'))
    
    try:
        # Check for existing user
        existing_user = User.query.filter(
            (User.nip == form_data['nip']) |
            (func.lower(User.username) == form_data['username'].lower()) |
            (func.lower(User.email) == form_data['email'])
        ).first()
        
        if existing_user:
            if existing_user.nip == form_data['nip']:
                flash('NIP sudah digunakan', 'error')
            elif existing_user.username.lower() == form_data['username'].lower():
                flash('Username sudah digunakan', 'error')
            else:
                flash('Email sudah terdaftar', 'error')
            return redirect(url_for('admin.manajemen_admin'))
        
        # Create new admin
        new_admin = User(
            nip=form_data['nip'],
            username=form_data['username'],
            full_name=form_data['full_name'],
            email=form_data['email'],
            password=generate_password_hash(form_data['password']),
            phone=form_data['phone'],
            role=UserRole.ADMIN,
            email_verified=False,
            created_by=current_user.id
        )
        
        db.session.add(new_admin)
        db.session.flush()  # Get the ID before commit
        
        # Log the action
        AuditLog.log(
            user_id=current_user.id,
            action='CREATE_ADMIN',
            model='User',
            model_id=new_admin.id,
            description=f"Created admin {form_data['username']}",
            ip_address=request.remote_addr
        )
        
        # Send admin-specific emails
        try:
            send_admin_verification_email(new_admin)
            send_welcome_email(new_admin, form_data['password'])
        except EmailSendError as e:
            current_app.logger.error(f"Email sending failed: {str(e)}")
            flash('Admin berhasil dibuat tetapi email tidak terkirim', 'warning')
        
        db.session.commit()
        
        flash(f'Admin {form_data["full_name"]} berhasil dibuat', 'success')
        return redirect(url_for('admin.manajemen_admin'))
        
    except IntegrityError as e:
        db.session.rollback()
        if "users.nip" in str(e.orig):
            flash('NIP sudah digunakan oleh admin lain', 'error')
        else:
            current_app.logger.error(f"Integrity error creating admin: {str(e)}")
            flash('Gagal membuat admin (data konflik)', 'error')
        return redirect(url_for('admin.manajemen_admin'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating admin: {str(e)}", exc_info=True)
        flash('Gagal membuat admin', 'error')
        return redirect(url_for('admin.manajemen_admin'))

@admin_bp.route('/superadmin/edit-admin/<int:admin_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def superadmin_edit_user(admin_id):
    """Edit admin user details"""
    admin = User.query.filter_by(id=admin_id, role=UserRole.ADMIN).first_or_404()
    
    if request.method == 'POST':
        return handle_admin_update(admin)
        
    return render_template('superadmin/edit_admin.html', admin=admin)

def handle_admin_update(admin):
    """Handle admin update form submission"""
    form_data = {
        'email': request.form.get('email', '').lower().strip(),
        'full_name': request.form.get('full_name', '').strip(),
        'phone': request.form.get('phone', '').strip(),
        'is_active': request.form.get('is_active', '0') == '1'
    }
    
    try:
        # Validate email uniqueness
        if User.query.filter(
            (User.email == form_data['email']) &
            (User.id != admin.id)
        ).first():
            flash('Email sudah digunakan oleh admin lain', 'error')
            return redirect(url_for('admin.superadmin_edit_user', admin_id=admin.id))
        
        # Update admin details
        admin.email = form_data['email']
        admin.full_name = form_data['full_name']
        admin.phone = form_data['phone']
        admin.is_active = form_data['is_active']
        
        # Log the action
        AuditLog.log(
            user_id=current_user.id,
            action='UPDATE_ADMIN',
            model='User',
            model_id=admin.id,
            description=f"Updated admin {admin.username}"
        )
        
        db.session.commit()
        flash('Data admin berhasil diperbarui', 'success')
        return redirect(url_for('admin.manajemen_admin'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating admin: {str(e)}", exc_info=True)
        flash('Gagal memperbarui data admin', 'error')
        return redirect(url_for('admin.superadmin_edit_user', admin_id=admin.id))


@admin_bp.route('/superadmin/delete-admin/<int:admin_id>', methods=['POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def superadmin_delete_user(admin_id):
    """Permanently delete admin user from database"""
    admin = User.query.filter_by(id=admin_id, role=UserRole.ADMIN).first_or_404()
    
    # Prevent self-deletion
    if admin.id == current_user.id:
        flash('Tidak dapat menghapus akun sendiri', 'error')
        return redirect(url_for('admin.manajemen_admin'))
    
    try:
        # Log the action before deletion
        AuditLog.log(
            user_id=current_user.id,
            action='PERMANENT_DELETE_ADMIN',
            model='User',
            model_id=admin.id,
            description=f"Permanently deleted admin {admin.username}",
            ip_address=request.remote_addr
        )
        
        # Get data for reference before deletion
        deleted_username = admin.username
        deleted_email = admin.email
        
        # Permanent delete
        db.session.delete(admin)
        db.session.commit()
        
        flash(f'Admin {deleted_username} ({deleted_email}) telah dihapus permanen dari sistem', 'success')
        return redirect(url_for('admin.manajemen_admin'))
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error permanently deleting admin: {str(e)}", exc_info=True)
        flash('Gagal menghapus admin secara permanen', 'error')
        return redirect(url_for('admin.manajemen_admin'))


@admin_bp.route('/superadmin/deactivate-admin/<int:admin_id>', methods=['POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def deactivate_admin(admin_id):
    admin = User.query.filter_by(id=admin_id, role=UserRole.ADMIN).first_or_404()
    
    if admin.id == current_user.id:
        flash('Tidak dapat menonaktifkan akun sendiri', 'error')
        return redirect(url_for('admin.manajemen_admin'))
    
    try:
        admin.is_active = False
        admin.deactivated_at = datetime.utcnow()
        admin.deactivated_by = current_user.id
        
        # Log audit
        AuditLog.log(
            user_id=current_user.id,
            action='DEACTIVATE_ADMIN',
            model='User',
            model_id=admin.id,
            description=f"Deactivated admin {admin.username}"
        )
        
        db.session.commit()
        
        # Kirim email notifikasi
        send_admin_status_email(admin, current_user, False)
        
        flash('Admin berhasil dinonaktifkan', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deactivating admin: {str(e)}")
        flash('Gagal menonaktifkan admin', 'error')
    
    return redirect(url_for('admin.manajemen_admin'))


@admin_bp.route('/superadmin/activate-admin/<int:admin_id>', methods=['POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def activate_admin(admin_id):
    admin = User.query.filter_by(id=admin_id, role=UserRole.ADMIN).first_or_404()
    
    try:
        admin.is_active = True
        admin.deactivated_at = None
        admin.deactivated_by = None
        
        # Log audit
        AuditLog.log(
            user_id=current_user.id,
            action='ACTIVATE_ADMIN',
            model='User',
            model_id=admin.id,
            description=f"Activated admin {admin.username}"
        )
        
        db.session.commit()
        
        # Kirim email notifikasi
        send_admin_status_email(admin, current_user, True)
        
        flash('Admin berhasil diaktifkan kembali', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error activating admin: {str(e)}")
        flash('Gagal mengaktifkan admin', 'error')
    
    return redirect(url_for('admin.manajemen_admin'))


@admin_bp.route('/superadmin/reset-admin-password/<int:admin_id>', methods=['POST'])
@login_required
@role_required(UserRole.SUPERADMIN)
def reset_admin_password(admin_id):
    """Reset admin password to default"""
    admin = User.query.filter_by(id=admin_id, role=UserRole.ADMIN).first_or_404()
    
    try:
        default_password = generate_random_password()  # Fungsi buat password acak
        admin.password = generate_password_hash(default_password)
        admin.must_change_password = True
        
        AuditLog.log(
            user_id=current_user.id,
            action='RESET_ADMIN_PASSWORD',
            model='User',
            model_id=admin.id,
            description=f"Reset password for admin {admin.username}"
        )
        
        db.session.commit()
        
        # Kirim email dengan password baru
        send_admin_password_reset_email(admin, default_password)
        
        flash('Password admin berhasil direset', 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error resetting password: {str(e)}")
        flash('Gagal mereset password admin', 'error')
    
    return redirect(url_for('admin.manajemen_admin'))


def generate_random_password(length=12):
    """Generate random password"""
    chars = string.ascii_letters + string.digits + '!@#$%^&*'
    return ''.join(random.choice(chars) for _ in range(length))


# Supporting functions
def validate_admin_form(form_data):
    errors = []
    
    if not form_data['nip']:
        errors.append('NIP wajib diisi')
    elif not form_data['nip'].isdigit():
        errors.append('NIP harus berupa angka')
    
    if not form_data['username']:
        errors.append('Username wajib diisi')
    elif len(form_data['username']) < 3:
        errors.append('Username minimal 3 karakter')
    
    if not form_data['full_name']:
        errors.append('Nama lengkap wajib diisi')
    
    if not form_data['email']:
        errors.append('Email wajib diisi')
    elif '@' not in form_data['email']:
        errors.append('Format email tidak valid')
    
    if not form_data['password']:
        errors.append('Password wajib diisi')
    elif len(form_data['password']) < 8:
        errors.append('Password minimal 8 karakter')
    
    return errors


def check_database_health():
    """Check database connection health"""
    try:
        db.session.execute(text('SELECT 1'))
        return {'status': 'healthy', 'message': 'Database connected'}
    except Exception as e:
        return {'status': 'unhealthy', 'message': str(e)}

def check_email_service():
    """Check email service availability"""
    try:
        with mail.connect() as conn:
            return {'status': 'healthy', 'message': 'Email service available'}
    except Exception as e:
        return {'status': 'unhealthy', 'message': str(e)}

# ===================== ADMIN ROUTES =====================
@admin_bp.route('/admin/dashboard')
@login_required
@role_required(UserRole.ADMIN)
def admin_dashboard():
    # Get basic statistics
    total_cuti = Cuti.query.count()
    pending_cuti = Cuti.query.filter_by(status=CutiStatus.PENDING).count()
    total_pegawai = User.query.filter_by(role=UserRole.PEGAWAI).count()
    
    # Get recent leave requests with proper model instances
    recent_requests = Cuti.query.options(
        db.joinedload(Cuti.pemohon)
    ).order_by(
        Cuti.created_at.desc()
    ).limit(5).all()
    
    # Get leave distribution by status
    status_distribution = db.session.query(
        Cuti.status,
        db.func.count(Cuti.id).label('count')
    ).group_by(Cuti.status).all()
    
    # Get monthly trend data (last 6 months)
    six_months_ago = datetime.now() - timedelta(days=180)
    
    monthly_trend = db.session.query(
        db.func.strftime('%Y-%m', Cuti.created_at).label('month'),
        db.func.count(Cuti.id).label('count')
    ).filter(
        Cuti.created_at >= six_months_ago
    ).group_by(
        db.func.strftime('%Y-%m', Cuti.created_at)
    ).order_by(
        db.func.strftime('%Y-%m', Cuti.created_at)
    ).all()
    
    # Format data for charts
    status_data = {
        'labels': [status[0].value for status in status_distribution],
        'counts': [status[1] for status in status_distribution]
    }
    
    trend_data = {
        'months': [trend[0] for trend in monthly_trend],
        'counts': [trend[1] for trend in monthly_trend]
    }
    
    return render_template('admin/dashboard.html',
        total_cuti=total_cuti,
        pending_cuti=pending_cuti,
        total_pegawai=total_pegawai,
        recent_requests=recent_requests,
        status_data=status_data,
        trend_data=trend_data,
        current_year=datetime.now().year
    )

@admin_bp.route('/update-supervisor', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def update_supervisor():
    pegawai_id = request.form.get('pegawai_id')
    atasan_id = request.form.get('atasan_id') or None  # Allow null
    
    pegawai = User.query.get_or_404(pegawai_id)
    pegawai.atasan_id = atasan_id
    
    try:
        db.session.commit()
        flash('Atasan berhasil diperbarui', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Gagal memperbarui atasan', 'error')
        current_app.logger.error(f"Error updating supervisor: {str(e)}")
    
    return redirect(url_for('admin.manajemen_pegawai'))

@admin_bp.route('/admin/manajemen-pegawai')
@login_required
@role_required(UserRole.ADMIN)
def manajemen_pegawai():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Query dengan pagination
    pegawai_query = User.query.filter_by(role=UserRole.PEGAWAI)
    pagination = pegawai_query.paginate(page=page, per_page=per_page, error_out=False)
    
    atasans = User.query.filter_by(role=UserRole.ATASAN).all()
    
    return render_template('admin/manajemen_pegawai.html', 
                         pegawai_list=pagination.items,
                         pagination=pagination,
                         atasans=atasans,
                         total_pegawai=pegawai_query.count())

@admin_bp.route('/admin/tambah-pegawai', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def tambah_pegawai():
    if request.method == 'POST':
        try:
            # Ambil data dari form
            username = request.form.get('username').strip()
            email = request.form.get('email').strip().lower()
            password = request.form.get('password')
            full_name = request.form.get('full_name').strip()
            nip = request.form.get('nip', '').strip()
            jabatan = request.form.get('jabatan', '').strip()
            golongan = request.form.get('golongan', '').strip()
            atasan_id = request.form.get('atasan_id', type=int)
            
            # Validasi sederhana
            if not all([username, email, password, full_name]):
                flash('Data wajib tidak lengkap', 'error')
                return redirect(url_for('admin.tambah_pegawai'))
            
            # Buat user baru
            new_user = User(
                username=username,
                email=email,
                password=generate_password_hash(password),
                full_name=full_name,
                nip=nip if nip else None,
                jabatan=jabatan if jabatan else None,
                golongan=golongan if golongan else None,
                atasan_id=atasan_id if atasan_id else None,
                role=UserRole.PEGAWAI,
                email_verified=True,
                is_active=True
            )
            
            db.session.add(new_user)
            db.session.commit()
            
            flash('Pegawai baru berhasil ditambahkan', 'success')
            return redirect(url_for('admin.manajemen_pegawai'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal menambahkan pegawai: {str(e)}', 'error')
            return redirect(url_for('admin.tambah_pegawai'))
    
    # GET request - tampilkan form
    atasans = User.query.filter_by(role=UserRole.ATASAN).all()
    return render_template('admin/tambah_pegawai.html', atasans=atasans)

@admin_bp.route('/admin/edit-pegawai/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_pegawai(user_id):
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        try:
            # Update data
            user.username = request.form.get('username', user.username).strip()
            user.email = request.form.get('email', user.email).strip().lower()
            user.full_name = request.form.get('full_name', user.full_name).strip()
            user.nip = request.form.get('nip', user.nip).strip()
            user.jabatan = request.form.get('jabatan', user.jabatan).strip()
            user.golongan = request.form.get('golongan', user.golongan).strip()
            user.atasan_id = request.form.get('atasan_id', type=int) or None
            
            # Jika ada password baru
            if request.form.get('new_password'):
                user.password = generate_password_hash(request.form.get('new_password'))
            
            db.session.commit()
            flash('Data pegawai berhasil diperbarui', 'success')
            return redirect(url_for('admin.manajemen_pegawai'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Gagal memperbarui data: {str(e)}', 'error')
            return redirect(url_for('admin.edit_pegawai', user_id=user_id))
    
    # GET request - tampilkan form edit
    atasans = User.query.filter_by(role=UserRole.ATASAN).all()
    return render_template('admin/edit_pegawai.html', 
                         user=user,
                         atasans=atasans)

@admin_bp.route('/admin/detail-pegawai/<int:user_id>')
@login_required
@role_required(UserRole.ADMIN)
def detail_pegawai(user_id):
    user = User.query.get_or_404(user_id)
    return render_template('admin/detail_pegawai.html', user=user)

@admin_bp.route('/admin/nonaktifkan-pegawai/<int:user_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def nonaktifkan_pegawai(user_id):
    user = User.query.get_or_404(user_id)
    
    try:
        user.is_active = False
        db.session.commit()
        flash(f'Akun {user.username} berhasil dinonaktifkan', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal menonaktifkan akun: {str(e)}', 'error')
    
    return redirect(url_for('admin.manajemen_pegawai'))

@admin_bp.route('/admin/aktifkan-pegawai/<int:user_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def aktifkan_pegawai(user_id):
    user = User.query.get_or_404(user_id)
    
    try:
        user.is_active = True
        db.session.commit()
        flash(f'Akun {user.username} berhasil diaktifkan', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Gagal mengaktifkan akun: {str(e)}', 'error')
    
    return redirect(url_for('admin.manajemen_pegawai'))

@admin_bp.route('/admin/manajemen-user')
@login_required
@role_required(UserRole.ADMIN)
def manajemen_user():
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    query = User.query
    
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) | 
            (User.email.ilike(f'%{search}%')) |
            (User.jabatan.ilike(f'%{search}%')))
    
    users = query.order_by(User.username).paginate(page=page, per_page=per_page)
    
    return render_template('admin/manajemen_user.html', 
                         users=users,
                         search=search)

# ============ PASSWORD RESET ============
@admin_bp.route('/reset-password-atasan/<int:atasan_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def reset_password_atasan(atasan_id):
    atasan = User.query.get_or_404(atasan_id)
    new_password = request.form.get('new_password')
    
    # Validate password
    if not new_password or len(new_password) < 8:
        flash('Password harus minimal 8 karakter', 'error')
        return redirect(url_for('admin.edit_atasan', atasan_id=atasan_id))
    
    try:
        # Password validation
        if not validate_password(new_password):
            flash('Password harus mengandung huruf besar, kecil, angka, dan spesial karakter', 'error')
            return redirect(url_for('admin.edit_atasan', atasan_id=atasan_id))
        
        atasan.password = generate_password_hash(new_password)
        atasan.must_change_password = True
        atasan.last_password_change = datetime.utcnow()
        
        # Audit log
        AuditLog.log(
            user_id=current_user.id,
            action='UPDATE',
            table_name='User',
            record_id=atasan.id,
            message=f'Password reset for atasan {atasan.username}'
        )
        
        db.session.commit()
        flash('Password atasan berhasil direset', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error resetting password for atasan {atasan_id}: {str(e)}", exc_info=True)
        flash('Gagal mereset password', 'error')
    
    return redirect(url_for('admin.edit_atasan', atasan_id=atasan_id))

# ============ ATASAN MANAGEMENT ============
@admin_bp.route('/manajemen-atasan')
@login_required
@role_required(UserRole.ADMIN)
def manajemen_atasan():
    atasans = User.query.filter_by(role=UserRole.ATASAN).order_by(User.full_name).all()
    return render_template('admin/manajemen_atasan.html', atasans=atasans)

@admin_bp.route('/tambah-atasan', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def tambah_atasan():
    if request.method == 'POST':
        form_data = request.form.to_dict()
        
        # Validasi
        errors = []
        if not form_data.get('nip'): errors.append('NIP wajib diisi')
        if not form_data.get('username'): errors.append('Username wajib diisi')
        if not form_data.get('email'): errors.append('Email wajib diisi')
        if not form_data.get('full_name'): errors.append('Nama lengkap wajib diisi')  # Changed from full_name to name
        if not form_data.get('jabatan'): errors.append('Jabatan wajib diisi')
        
        # Simplified email validation
        email = form_data.get('email', '').strip()
        if '@' not in email or '.' not in email.split('@')[-1]:
            errors.append('Format email tidak valid')
            
        if 'password' in form_data and form_data['password']:
            if len(form_data['password']) < 8:
                errors.append('Password minimal 8 karakter')
            if form_data['password'] != form_data.get('confirm_password', ''):
                errors.append('Konfirmasi password tidak sama')
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('admin/tambah_atasan.html', form=form_data)
        
        try:
            # Cek duplikat
            existing = User.query.filter(
                (User.nip == form_data['nip']) |
                (User.username == form_data['username']) |
                (User.email == form_data['email'])
            ).first()
            
            if existing:
                flash('NIP/Username/Email sudah terdaftar', 'error')
                return render_template('admin/tambah_atasan.html', form=form_data)
            
            # Buat atasan baru
            atasan = User(
                nip=form_data['nip'],
                username=form_data['username'],
                email=form_data['email'],
                password=generate_password_hash(form_data['password']),
                full_name=form_data['full_name'],  # Changed to match form field name
                jabatan=form_data['jabatan'],
                golongan=form_data.get('golongan', ''),
                role=UserRole.ATASAN,
                email_verified=True,
                created_by=current_user.id
            )
            
            db.session.add(atasan)
            db.session.commit()
            
            flash('Atasan berhasil ditambahkan', 'success')
            return redirect(url_for('admin.manajemen_atasan'))
            
        except Exception as e:
            db.session.rollback()
            flash('Terjadi kesalahan saat menyimpan data', 'error')
            current_app.logger.error(f"Error adding atasan: {str(e)}")
            return render_template('admin/tambah_atasan.html', form=form_data)
    
    return render_template('admin/tambah_atasan.html')

@admin_bp.route('/edit-atasan/<int:atasan_id>', methods=['GET', 'POST'])
@login_required
@role_required(UserRole.ADMIN)
def edit_atasan(atasan_id):
    atasan = User.query.get_or_404(atasan_id)
    
    if request.method == 'POST':
        form_data = request.form.to_dict()
        
        # Validasi
        errors = []
        if not form_data.get('nip'): errors.append('NIP wajib diisi')
        if not form_data.get('username'): errors.append('Username wajib diisi')
        if not form_data.get('email'): errors.append('Email wajib diisi')
        if not form_data.get('full_name'): errors.append('Nama lengkap wajib diisi')
        if not form_data.get('jabatan'): errors.append('Jabatan wajib diisi')
        
        # Simplified email validation
        email = form_data.get('email', '').strip()
        if '@' not in email or '.' not in email.split('@')[-1]:
            errors.append('Format email tidak valid')
            
        if errors:
            for error in errors:
                flash(error, 'error')
            return render_template('admin/edit_atasan.html', atasan=atasan)
        
        try:
            # Cek duplikat (kecuali diri sendiri)
            existing = User.query.filter(
                (User.id != atasan_id) & (
                    (User.nip == form_data['nip']) |
                    (User.username == form_data['username']) |
                    (User.email == form_data['email'])
                )
            ).first()
            
            if existing:
                flash('NIP/Username/Email sudah terdaftar', 'error')
                return render_template('admin/edit_atasan.html', atasan=atasan)
            
            # Update data
            atasan.nip = form_data['nip']
            atasan.username = form_data['username']
            atasan.email = form_data['email']
            atasan.full_name = form_data['full_name']
            atasan.jabatan = form_data['jabatan']
            atasan.golongan = form_data.get('golongan', '')
            
            # Update password jika diisi
            if form_data.get('password'):
                atasan.password = generate_password_hash(form_data['password'])
            
            db.session.commit()
            
            flash('Data atasan berhasil diperbarui', 'success')
            return redirect(url_for('admin.manajemen_atasan'))
            
        except Exception as e:
            db.session.rollback()
            flash('Terjadi kesalahan saat menyimpan data', 'error')
            current_app.logger.error(f"Error updating atasan: {str(e)}")
    
    return render_template('admin/edit_atasan.html', atasan=atasan)

@admin_bp.route('/hapus-atasan/<int:atasan_id>', methods=['POST'])
@login_required
@role_required(UserRole.ADMIN)
def hapus_atasan(atasan_id):
    atasan = User.query.get_or_404(atasan_id)
    
    try:
        # Cek apakah atasan memiliki bawahan
        if atasan.subordinates.count() > 0:
            flash('Tidak dapat menghapus atasan yang masih memiliki bawahan', 'error')
            return redirect(url_for('admin.manajemen_atasan'))
        
        db.session.delete(atasan)
        db.session.commit()
        
        flash('Atasan berhasil dihapus', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Terjadi kesalahan saat menghapus atasan', 'error')
        current_app.logger.error(f"Error deleting atasan: {str(e)}")
    
    return redirect(url_for('admin.manajemen_atasan'))

# ===================== UTILITY ROUTES =====================
@admin_bp.route('/admin/export-users')
@login_required
@role_required(UserRole.ADMIN, UserRole.SUPERADMIN)
def export_users():
    users = User.query.all()
    
    # Create DataFrame
    data = []
    for user in users:
        data.append({
            'NIP': user.nip,
            'Username': user.username,
            'Email': user.email,
            'Role': user.role.value,
            'Jabatan': user.jabatan,
            'Status': 'Aktif' if user.is_active else 'Nonaktif',
            'Tanggal Dibuat': user.created_at.strftime('%Y-%m-%d')
        })
    
    df = pd.DataFrame(data)
    
    # Export to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Users', index=False)
    
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
    return response

@admin_bp.route('/laporan-cuti', methods=['GET'])
@login_required
@role_required(UserRole.ADMIN)
def laporan_cuti():
    try:
        # 1. Ambil parameter filter dengan validasi
        tahun = request.args.get('tahun', default=datetime.now().year, type=int)
        bulan = request.args.get('bulan', type=int)
        status = request.args.get('status')
        jenis_cuti = request.args.get('jenis_cuti')
        
        # 2. Buat query dasar dengan relasi yang sudah ada di model
        query = Cuti.query.options(
            db.joinedload(Cuti.pemohon),  # Menggunakan relasi pemohon yang sudah didefinisikan
            db.joinedload(Cuti.penyetuju)  # Menggunakan relasi penyetuju yang sudah didefinisikan
        )
        
        # 3. Terapkan filter
        # Filter tahun wajib
        query = query.filter(
            db.extract('year', Cuti.tanggal_mulai) == tahun
        )
        
        # Filter bulan opsional
        if bulan and 1 <= bulan <= 12:
            query = query.filter(
                db.extract('month', Cuti.tanggal_mulai) == bulan
            )
        
        # Filter status opsional
        if status and status in [s.value for s in CutiStatus]:
            query = query.filter(Cuti.status == status)
            
        # Filter jenis cuti opsional
        if jenis_cuti and jenis_cuti in [j.value for j in JenisCuti]:
            query = query.filter(Cuti.jenis_cuti == jenis_cuti)

        # 4. Paginasi data
        page = request.args.get('page', 1, type=int)
        per_page = 20
        paginated_data = query.order_by(
            Cuti.tanggal_mulai.desc()
        ).paginate(
            page=page, 
            per_page=per_page,
            error_out=False
        )

        # 5. Ambil opsi filter tahun (sama seperti di dashboard)
        tahun_options = db.session.query(
            db.extract('year', Cuti.tanggal_mulai).label('tahun')
        ).distinct().order_by(
            db.desc('tahun')
        ).all()
        
        # 6. Hitung statistik ringkasan (format sama dengan dashboard)
        summary = {
            'total': query.count(),
            'approved': query.filter(Cuti.status == CutiStatus.APPROVED).count(),
            'pending': query.filter(Cuti.status == CutiStatus.PENDING).count(),
            'rejected': query.filter(Cuti.status == CutiStatus.REJECTED).count()
        }

        # 7. Render template dengan struktur data yang konsisten
        return render_template(
            'admin/laporan_cuti.html',
            cuti_data=paginated_data,
            tahun=tahun,
            bulan=bulan,
            status=status,
            jenis_cuti=jenis_cuti,
            tahun_options=[t[0] for t in tahun_options],
            summary=summary,
            CutiStatus=CutiStatus,
            JenisCuti=JenisCuti
        )

    except ValueError as e:
        current_app.logger.error(f"Parameter tidak valid di laporan_cuti: {str(e)}")
        flash('Parameter filter tidak valid', 'error')
        return redirect(url_for('admin.laporan_cuti', tahun=datetime.now().year))
        
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Error database di laporan_cuti: {str(e)}")
        flash('Terjadi kesalahan database', 'error')
        return redirect(url_for('admin.admin_dashboard'))
        
    except Exception as e:
        current_app.logger.error(f"Error tidak terduga di laporan_cuti: {str(e)}")
        flash('Terjadi kesalahan sistem', 'error')
        return redirect(url_for('admin.admin_dashboard'))

def get_leave_summary(tahun, bulan=None, departemen=None):
    """Helper function to generate leave summary statistics"""
    query = db.session.query(Cuti).join(User)
    
    query = query.filter(extract('year', Cuti.tanggal_mulai) == tahun)
    if bulan:
        query = query.filter(extract('month', Cuti.tanggal_mulai) == bulan)
    
    total = query.count()
    
    status_counts = dict(db.session.query(
        Cuti.status,
        func.count(Cuti.id)
    ).group_by(Cuti.status).all())
    
    jenis_counts = dict(db.session.query(
        Cuti.jenis_cuti,
        func.count(Cuti.id)
    ).group_by(Cuti.jenis_cuti).all())
    
    return {
        'total': total,
        'status': {
            'APPROVED': status_counts.get(CutiStatus.APPROVED, 0),
            'REJECTED': status_counts.get(CutiStatus.REJECTED, 0),
            'PENDING': status_counts.get(CutiStatus.PENDING, 0)
        },
        'jenis': {
            'TAHUNAN': jenis_counts.get(JenisCuti.TAHUNAN, 0),
            'SAKIT': jenis_counts.get(JenisCuti.SAKIT, 0),
            'MELAHIRKAN': jenis_counts.get(JenisCuti.MELAHIRKAN, 0),
            'BESAR': jenis_counts.get(JenisCuti.BESAR, 0)
        }
    }


@admin_bp.route('/cetak-surat-cuti/<int:cuti_id>', methods=['GET'])
@login_required
@role_required(UserRole.ADMIN)
def cetak_surat_cuti(cuti_id):
    # Query with all necessary joins
    cuti = db.session.query(
        Cuti,
        User.full_name,
        User.nip,
        User.jabatan,
        User.jenis_kelamin
    ).join(
        User, Cuti.user_id == User.id
    ).filter(
        Cuti.id == cuti_id
    ).first_or_404()

    # Format dates
    tanggal_format = datetime.now().strftime("%d %B %Y")
    tahun = datetime.now().year
    
    return render_template('admin/cetak_surat.html',
        cuti={
            'id': cuti[0].id,
            'jenis_cuti': cuti[0].jenis_cuti.value if isinstance(cuti[0].jenis_cuti, Enum) else cuti[0].jenis_cuti,
            'tanggal_mulai': cuti[0].tanggal_mulai.strftime("%d %B %Y"),
            'tanggal_selesai': cuti[0].tanggal_selesai.strftime("%d %B %Y"),
            'jumlah_hari': cuti[0].jumlah_hari,
            'username': cuti.full_name,
            'nip': cuti.nip or '-',
            'jabatan': cuti.jabatan or '-',
            'jenis_kelamin': cuti.jenis_kelamin,
            'tahun': tahun,
            'tanggal_format': tanggal_format,
            'perihal_cuti': cuti[0].perihal_cuti
        }
    )



@admin_bp.route('/cetak-laporan-cuti', methods=['GET'])
@login_required
@role_required(UserRole.ADMIN)
def cetak_laporan_cuti():
    try:
        # 1. Validate parameters
        tahun = request.args.get('tahun', default=datetime.now().year, type=int)
        bulan = request.args.get('bulan', type=int)
        status = request.args.get('status')
        jenis_cuti = request.args.get('jenis_cuti')
        format_file = request.args.get('format', 'pdf').lower()
        orientation = request.args.get('orientation', 'portrait')
        
        if format_file not in ['pdf', 'excel']:
            flash('Format file tidak valid. Pilih PDF atau Excel.', 'error')
            return redirect(url_for('admin.laporan_cuti'))

        # 2. Build query
        query = Cuti.query.options(
            db.joinedload(Cuti.pemohon),
            db.joinedload(Cuti.penyetuju)
        ).filter(
            db.extract('year', Cuti.tanggal_mulai) == tahun
        )
        
        if bulan and 1 <= bulan <= 12:
            query = query.filter(
                db.extract('month', Cuti.tanggal_mulai) == bulan
            )
        
        if status and status in [s.value for s in CutiStatus]:
            query = query.filter(Cuti.status == status)
            
        if jenis_cuti and jenis_cuti in [j.value for j in JenisCuti]:
            query = query.filter(Cuti.jenis_cuti == jenis_cuti)

        # 3. Get all data
        data_cuti = query.order_by(Cuti.tanggal_mulai).all()
        
        if not data_cuti:
            flash('Tidak ada data cuti untuk diekspor', 'warning')
            return redirect(url_for('admin.laporan_cuti'))

        # 4. Prepare export data
        export_data = []
        for cuti in data_cuti:
            export_data.append({
                'NIP': cuti.pemohon.nip if cuti.pemohon else '-',
                'Nama Pegawai': cuti.pemohon.full_name if cuti.pemohon else '-',
                'Jenis Cuti': cuti.jenis_cuti.value,
                'Tanggal Mulai': cuti.tanggal_mulai.strftime('%d/%m/%Y'),
                'Tanggal Selesai': cuti.tanggal_selesai.strftime('%d/%m/%Y'),
                'Lama Cuti': f"{cuti.jumlah_hari} hari",
                'Status': cuti.status.value,
                'Disetujui Oleh': cuti.penyetuju.full_name if cuti.penyetuju else '-',
                'Tanggal Pengajuan': cuti.created_at.strftime('%d/%m/%Y %H:%M'),
                'Perihal': cuti.perihal_cuti or '-',
                'Alasan Penolakan': cuti.alasan_penolakan or '-'
            })

        # 5. Generate file
        if format_file == 'excel':
            return generate_excel_report(export_data, tahun, bulan)
        else:
            pdf_buffer = generate_pdf_report(export_data, tahun, bulan, orientation)
            
            # Membuat nama file yang lebih baik
            bulan_text = f"_bulan_{bulan}" if bulan else ""
            filename = f"laporan_cuti_{tahun}{bulan_text}.pdf"
            
            # Return response dengan header yang benar
            response = make_response(pdf_buffer.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response

    except ValueError as e:
        current_app.logger.error(f"Invalid parameter: {str(e)}")
        flash('Parameter filter tidak valid', 'error')
        return redirect(url_for('admin.laporan_cuti'))
        
    except Exception as e:
        current_app.logger.error(f"Export error: {str(e)}", exc_info=True)
        flash('Gagal membuat laporan ekspor', 'error')
        return redirect(url_for('admin.laporan_cuti'))


from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from flask import current_app
from io import BytesIO

def register_fonts():
    try:
        # Path ke folder fonts - gunakan path absolut untuk memastikan
        font_dir = os.path.join(current_app.root_path, 'static', 'fonts')
        
        # Daftarkan font yang benar-benar ada
        # Dari file yang Anda tunjukkan, kita punya:
        # - Nunito-Regular.ttf
        # - Nunito-Bold.ttf
        # - Nunito-Italic.ttf
        # - Nunito-BoldItalic.ttf
        
        # Pastikan path file benar
        regular_path = os.path.join(font_dir, 'Nunito-Regular.ttf')
        bold_path = os.path.join(font_dir, 'Nunito-Bold.ttf')
        italic_path = os.path.join(font_dir, 'Nunito-Italic.ttf')
        bold_italic_path = os.path.join(font_dir, 'Nunito-BoldItalic.ttf')
        
        # Verifikasi file ada sebelum mendaftarkan
        if not os.path.exists(regular_path):
            raise FileNotFoundError(f"Font file not found: {regular_path}")
        
        pdfmetrics.registerFont(TTFont('Nunito', regular_path))
        pdfmetrics.registerFont(TTFont('Nunito-Bold', bold_path))
        pdfmetrics.registerFont(TTFont('Nunito-Italic', italic_path))
        pdfmetrics.registerFont(TTFont('Nunito-BoldItalic', bold_italic_path))
        
        return True
    except Exception as e:
        current_app.logger.error(f"Font registration error: {str(e)}")
        # Fallback ke font default jika gagal
        try:
            pdfmetrics.registerFont(TTFont('Nunito', 'Helvetica'))
            pdfmetrics.registerFont(TTFont('Nunito-Bold', 'Helvetica-Bold'))
            current_app.logger.warning("Using fallback Helvetica fonts")
            return False
        except:
            current_app.logger.error("Failed to register fallback fonts")
            raise

def generate_pdf_report(data, tahun, bulan, orientation='portrait', page_size='a4'):
    try:
        # 1. Register fonts
        current_app.logger.info("Registering fonts...")
        font_loaded = register_fonts()
        current_app.logger.info(f"Font registration {'successful' if font_loaded else 'used fallback fonts'}")
        
        styles = getSampleStyleSheet()
        
        # 2. Create custom styles
        styles.add(ParagraphStyle(
            name='DataStyle',
            fontName='Nunito',
            fontSize=8,
            leading=9,
            wordWrap='LTR'
        ))
        
        styles.add(ParagraphStyle(
            name='SmallDataStyle',
            parent=styles['DataStyle'],
            fontSize=7,
            leading=8
        ))

        # 3. Create document
        buffer = BytesIO()
        page_size = A4 if page_size.lower() == 'a4' else letter
        if orientation == 'landscape':
            page_size = landscape(page_size)
            
        doc = SimpleDocTemplate(buffer, 
                              pagesize=page_size,
                              leftMargin=0.4*inch,
                              rightMargin=0.4*inch,
                              topMargin=0.5*inch,
                              bottomMargin=0.5*inch)

        elements = []
        
        # 4. Report title
        title_style = styles['Title']
        title_style.fontName = 'Nunito-Bold'
        title = Paragraph(
            f"LAPORAN CUTI TAHUN {tahun} {'BULAN '+str(bulan) if bulan else 'SEMUA BULAN'}",
            title_style
        )
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        if not data:
            no_data_style = styles['BodyText']
            no_data_style.fontName = 'Nunito'
            elements.append(Paragraph("Tidak ada data cuti", no_data_style))
        else:
            # 5. Define all columns
            headers = [
                "No", 
                "NIP",
                "Nama Pegawai", 
                "Jenis Cuti", 
                "Tanggal Mulai",
                "Tanggal Selesai",
                "Lama Cuti", 
                "Status", 
                "Disetujui Oleh",
                "Tanggal Pengajuan",
                "Perihal",
                "Alasan Penolakan"
            ]
            
            # 6. Column widths (adjusted for landscape)
            col_widths = [
                0.3*inch,  # No
                0.8*inch,  # NIP
                1.2*inch,  # Nama Pegawai
                0.9*inch,  # Jenis Cuti
                0.7*inch,  # Tanggal Mulai
                0.7*inch,  # Tanggal Selesai
                0.5*inch,  # Lama Cuti
                0.6*inch,  # Status
                1.0*inch,  # Disetujui Oleh
                0.9*inch,  # Tanggal Pengajuan
                1.5*inch,  # Perihal
                1.5*inch   # Alasan Penolakan
            ]
            
            # 7. Build table data
            table_data = []
            
            # Header row style
            header_style = ParagraphStyle(
                name='HeaderStyle',
                parent=styles['Normal'],
                fontName='Nunito-Bold',
                fontSize=8,
                alignment=1,  # Center
                textColor=colors.white
            )
            table_data.append([Paragraph(h, header_style) for h in headers])
            
            # Data rows
            for i, row in enumerate(data, 1):
                table_data.append([
                    str(i),
                    Paragraph(row.get('NIP', '-'), styles['DataStyle']),
                    Paragraph(row.get('Nama Pegawai', '-'), styles['DataStyle']),
                    Paragraph(row.get('Jenis Cuti', '-'), styles['DataStyle']),
                    Paragraph(row.get('Tanggal Mulai', '-'), styles['DataStyle']),
                    Paragraph(row.get('Tanggal Selesai', '-'), styles['DataStyle']),
                    Paragraph(row.get('Lama Cuti', '-'), styles['DataStyle']),
                    Paragraph(row.get('Status', '-'), styles['DataStyle']),
                    Paragraph(row.get('Disetujui Oleh', '-'), styles['DataStyle']),
                    Paragraph(row.get('Tanggal Pengajuan', '-'), styles['DataStyle']),
                    Paragraph(row.get('Perihal', '-'), styles['SmallDataStyle']),
                    Paragraph(row.get('Alasan Penolakan', '-'), styles['SmallDataStyle'])
                ])
            
            # 8. Create table
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4e73df')),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('FONTNAME', (0,0), (-1,0), 'Nunito-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 8),
                ('FONTSIZE', (0,1), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('LEADING', (0,1), (-1,-1), 8),
            ]))
            
            elements.append(table)
        
        # 9. Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Optional: Save debug copy
        debug_path = os.path.join(current_app.root_path, 'debug_pdf.pdf')
        with open(debug_path, 'wb') as f:
            f.write(buffer.getvalue())
        current_app.logger.info(f"Debug PDF saved to: {debug_path}")
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        current_app.logger.error(f"PDF Generation Error: {str(e)}", exc_info=True)
        raise
        

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel_report(data, tahun, bulan):
    try:
        # Create a workbook and add worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Cuti"
        
        # Add title
        bulan_text = f"Bulan {bulan}" if bulan else "Semua Bulan"
        title = f"Laporan Cuti Tahun {tahun} {bulan_text}"
        ws.append([title])
        ws.merge_cells('A1:K1')  # Sesuaikan dengan jumlah kolom Anda
        title_cell = ws['A1']
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        if data:
            headers = list(data[0].keys())
            ws.append(headers)
            
            # Style headers
            header_row = ws[2]  # Baris kedua adalah header
            for cell in header_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4e73df", end_color="4e73df", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(bottom=Side(border_style="thin"))
        
            # Add data rows
            for row in data:
                ws.append([row[header] for header in headers])
                
            # Auto-adjust column widths
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                # Skip merged cells in the first row
                if col_idx == 1 and ws.cell(row=1, column=1).value == title:
                    continue
                    
                for cell in column:
                    try:
                        if cell.row > 1:  # Skip title row
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width > 10 else 10
            
            # Style data rows
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(border_style="thin"),
                                      right=Side(border_style="thin"),
                                      top=Side(border_style="thin"),
                                      bottom=Side(border_style="thin"))
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"laporan_cuti_{tahun}_{bulan or 'all'}.xlsx"
        return send_file(buffer, 
                       as_attachment=True, 
                       download_name=filename,
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        current_app.logger.error(f"Excel generation error: {str(e)}", exc_info=True)
        raise


def prepare_export_data(results):
    """Prepare data for export in consistent format"""
    return [{
        'NIP': user.nip,
        'Nama': user.full_name,
        'Jenis Cuti': cuti.jenis_cuti.value,
        'Tanggal Mulai': cuti.tanggal_mulai.strftime('%d/%m/%Y'),
        'Tanggal Selesai': cuti.tanggal_selesai.strftime('%d/%m/%Y'),
        'Durasi': cuti.jumlah_hari,
        'Status': cuti.status.value,
        'Pengajuan': cuti.created_at.strftime('%d/%m/%Y %H:%M'),
        'Atasan': get_supervisor_name(cuti.atasan_id),
        'Perihal': cuti.perihal_cuti
    } for cuti, user in results]

def get_supervisor_name(atasan_id):
    """Helper to get supervisor name with caching"""
    if not atasan_id:
        return '-'
    supervisor = User.query.get(atasan_id)
    return supervisor.full_name if supervisor else '-'

# ------------------------------------------------------

